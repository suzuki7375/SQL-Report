# -*- coding: utf-8 -*-

import argparse
import datetime
import os
import sys
import time

import pandas as pd
from pandas.api.types import is_datetime64_any_dtype
import pyodbc
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

SERVER = "omddb"
USERNAME = "PE_ReadOnlyUser"
PASSWORD = "pe@0505"
DRIVER = "ODBC Driver 17 for SQL Server"
DATABASE = "MEQueryManufacturingDatabase"

TARGET_OBJECT = "[工八_PEREADONLY].[LK2MES-DB-REAL].[dbo].[V_PE_PRD_TestResult_800G_BER_Symbol_Error_Test]"
BER_SQL_FILE = "BER_Symbol_Error_Test.sql"
MASTER_SQL_FILE = "MASTER.sql"

# 先看資料用
PREVIEW_N = 20

# 匯出用（先小量，確定 OK 再加大）
EXPORT_N = 2000

OUTPUT_EXTENSION = ".xlsx"
TEST_ITEM_HEADER = "測試項目"
CH_NUMBER_HEADER = "CHNumber"
COMPONENT_ID_HEADER = "COMPONENTID"
TESTNUMBER_HEADER = "TESTNUMBER"
DATA_ANALYSIS_SHEET = "Data Analysis"
ERROR_CODE_SHEET = "Error Code"
FUNCTION_TEMPLATE = "Function.xlsx"
ERROR_CODE_HEADER = "Error code"
FAILURE_CODE_HEADER = "FailureCodeID"
PARETO_COLUMNS = ["Error Code", "Fail Q'ty", "Failed Rate", "Cum%"]
PARETO_LIMIT = 10
STATION_ORDER = [
    "DDMI",
    "RT",
    "LT",
    "HT",
    "Burn In",
    "TC BER",
    "ATS",
    "Switch",
]
THREE_T_BER_ITEMS = [
    "1_Pretest",
    "2_Pretest",
    "3_Pretest",
    "4_Pretest",
    "5_Pretest",
    "6_Pretest",
    "7_Pretest",
    "8_Pretest",
    "1_RT",
    "2_RT",
    "3_RT",
    "4_RT",
    "5_RT",
    "6_RT",
    "7_RT",
    "8_RT",
    "1_LT",
    "2_LT",
    "3_LT",
    "4_LT",
    "5_LT",
    "6_LT",
    "7_LT",
    "8_LT",
    "1_HT",
    "2_HT",
    "3_HT",
    "4_HT",
    "5_HT",
    "6_HT",
    "7_HT",
    "8_HT",
]
THREE_T_BER_ITEM_SET = {item.lower() for item in THREE_T_BER_ITEMS}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--start-date", required=True, help="YYYY-MM-DD")
    parser.add_argument("--end-date", required=True, help="YYYY-MM-DD")
    parser.add_argument("--output-dir", help="輸出資料夾")
    return parser.parse_args()


def parse_date(value: str) -> datetime.date:
    try:
        return datetime.date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError("日期格式需為 YYYY-MM-DD") from exc


def conn_str_no_db() -> str:
    return (
        f"DRIVER={{{DRIVER}}};"
        f"SERVER={SERVER};"
        f"UID={USERNAME};"
        f"PWD={PASSWORD};"
        "TrustServerCertificate=yes;"
    )


def conn_str_with_db(db: str) -> str:
    return (
        f"DRIVER={{{DRIVER}}};"
        f"SERVER={SERVER};"
        f"DATABASE={db};"
        f"UID={USERNAME};"
        f"PWD={PASSWORD};"
        "TrustServerCertificate=yes;"
    )


def test_login() -> None:
    conn = pyodbc.connect(conn_str_no_db(), timeout=5)
    cur = conn.cursor()
    cur.execute("SELECT SYSTEM_USER, SUSER_SNAME(), @@SERVERNAME")
    print("✅ 登入成功（未指定 DATABASE）")
    print("🔎 Login info:", cur.fetchone())
    conn.close()


def load_sql(path: str) -> str:
    with open(path, "r", encoding="utf-8") as file:
        lines = file.read().splitlines()
    cleaned_lines = [line for line in lines if line.strip().upper() != "GO"]
    return "\n".join(cleaned_lines).strip()


def build_sorted_query(limit: int) -> str:
    return f"""
WITH base AS (
    SELECT *,
        TRY_CONVERT(date, SUBSTRING(TESTNUMBER, 2, 8)) AS test_date,
        DATETIMEFROMPARTS(
            TRY_CONVERT(int, SUBSTRING(TESTNUMBER, 2, 4)),
            TRY_CONVERT(int, SUBSTRING(TESTNUMBER, 6, 2)),
            TRY_CONVERT(int, SUBSTRING(TESTNUMBER, 8, 2)),
            TRY_CONVERT(int, SUBSTRING(TESTNUMBER, 10, 2)),
            TRY_CONVERT(int, SUBSTRING(TESTNUMBER, 12, 2)),
            0,
            0
        ) AS test_datetime
    FROM {TARGET_OBJECT}
)
SELECT TOP {limit} *
FROM base
WHERE test_date BETWEEN ? AND ?
ORDER BY test_datetime;
""".strip()


def build_export_query(limit: int, base_dir: str) -> str:
    sql_path = os.path.join(base_dir, BER_SQL_FILE)
    if not os.path.exists(sql_path):
        return build_sorted_query(limit)
    sql_text = load_sql(sql_path)
    return f"""
SELECT TOP {limit} *
FROM ({sql_text}) AS base
WHERE TRY_CONVERT(date, SUBSTRING(TESTNUMBER, 2, 8)) BETWEEN ? AND ?
ORDER BY
    DATETIMEFROMPARTS(
        TRY_CONVERT(int, SUBSTRING(TESTNUMBER, 2, 4)),
        TRY_CONVERT(int, SUBSTRING(TESTNUMBER, 6, 2)),
        TRY_CONVERT(int, SUBSTRING(TESTNUMBER, 8, 2)),
        TRY_CONVERT(int, SUBSTRING(TESTNUMBER, 10, 2)),
        TRY_CONVERT(int, SUBSTRING(TESTNUMBER, 12, 2)),
        0,
        0
    );
""".strip()


def apply_header(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    columns = list(df.columns)
    columns[0] = TEST_ITEM_HEADER
    df.columns = columns
    return df


def format_date_range(start_date: str, end_date: str) -> str:
    if start_date == end_date:
        return start_date
    return f"{start_date}_{end_date}"


def build_output_path(base_dir: str, start_date: str, end_date: str) -> str:
    base_name = os.path.splitext(os.path.basename(__file__))[0]
    date_range = format_date_range(start_date, end_date)
    filename = f"{base_name}_{date_range}{OUTPUT_EXTENSION}"
    return os.path.join(base_dir, filename)


def normalize_output_dir(base_dir: str, output_dir: str | None) -> str:
    if not output_dir:
        return base_dir
    expanded_dir = os.path.expanduser(output_dir)
    if not os.path.isabs(expanded_dir):
        return os.path.join(base_dir, expanded_dir)
    return expanded_dir


def ensure_unique_output_path(path: str) -> str:
    if not os.path.exists(path):
        return path
    root, ext = os.path.splitext(path)
    counter = 1
    while True:
        candidate = f"{root}_{counter}{ext}"
        if not os.path.exists(candidate):
            return candidate
        counter += 1


def classify_ch_number(value: str) -> str:
    text = str(value) if value is not None else ""
    if is_three_t_ber_channel(text):
        return "TC_BER"
    if is_tc_ber_channel(text):
        return "TC_BER"
    if "ATS" in text:
        return "ATS"
    if "DDMI" in text:
        return "DDMI"
    if "TP2TP3_LT" in text or "_LT" in text:
        return "LT"
    if "TP2TP3_HT" in text or "_HT" in text:
        return "HT"
    if "TP2TP3_RT" in text or "_RT" in text:
        return "RT"
    return "其他"


def find_column(columns: list[str], candidates: list[str]) -> str | None:
    lowered = {col.lower(): col for col in columns}
    for name in candidates:
        match = lowered.get(name.lower())
        if match:
            return match
    for col in columns:
        lowered_col = col.lower()
        if any(token in lowered_col for token in candidates):
            return col
    return None


def normalize_testnumber(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def find_component_column(columns: list[str]) -> str:
    column = find_column(
        columns,
        [COMPONENT_ID_HEADER.lower(), "component_id", "component"],
    )
    if not column:
        raise KeyError(f"查無欄位 {COMPONENT_ID_HEADER}")
    return column


def find_testnumber_column(columns: list[str]) -> str | None:
    return find_column(
        columns,
        [
            TESTNUMBER_HEADER.lower(),
            "test_number",
            "testnumber",
            "testno",
            "test_no",
        ],
    )


def build_date_series(df: pd.DataFrame) -> pd.Series | None:
    date_columns = [
        "test_date",
        "test_datetime",
        "TEST_DATETIME",
        "TestDateTime",
        "TEST_TIME",
        "TESTDATE",
    ]
    for column in date_columns:
        if column in df.columns:
            series = pd.to_datetime(df[column], errors="coerce")
            if series.notna().any():
                return series.dt.date

    for column in df.columns:
        if is_datetime64_any_dtype(df[column]):
            series = pd.to_datetime(df[column], errors="coerce")
            if series.notna().any():
                return series.dt.date

    candidate_columns = [
        column
        for column in df.columns
        if "date" in column.lower() or "time" in column.lower()
    ]
    for column in candidate_columns:
        series = pd.to_datetime(df[column], errors="coerce")
        if series.notna().any():
            return series.dt.date

    testnumber_column = find_testnumber_column(list(df.columns))
    if not testnumber_column:
        return None
    series = pd.to_datetime(
        df[testnumber_column].astype(str).str.slice(1, 9),
        format="%Y%m%d",
        errors="coerce",
    )
    if series.notna().any():
        return series.dt.date
    return None


def filter_df_by_date_range(
    df: pd.DataFrame,
    start_date: datetime.date,
    end_date: datetime.date,
) -> pd.DataFrame:
    if df.empty:
        return df
    date_series = build_date_series(df)
    if date_series is None:
        return df
    mask = (date_series >= start_date) & (date_series <= end_date)
    return df.loc[mask].copy()


def find_equipment_column(columns: list[str]) -> str | None:
    return find_column(
        columns,
        [
            "equipment",
            "equpment",
            "equupment",
            "equp",
            "eqp",
        ],
    )


def find_failure_code_column(columns: list[str]) -> str | None:
    return find_column(
        columns,
        [
            FAILURE_CODE_HEADER.lower(),
            "failure_code_id",
            "failure_code",
            "failurecodeid",
            "failurecode",
        ],
    )


def find_station_column(columns: list[str]) -> str | None:
    return find_column(columns, ["station", "teststation", "test_station"])


def find_result_column(columns: list[str]) -> str | None:
    return find_column(
        columns,
        [
            "result",
            "testresult",
            "test_result",
            "pass_fail",
            "passfail",
            "pf",
            "status",
        ],
    )


def find_ch_pass_fail_columns(columns: list[str]) -> list[str]:
    return [col for col in columns if "ch_pass_fail" in col.lower()]


def normalize_station(text: str) -> str | None:
    if not text:
        return None
    upper = text.upper()
    if is_three_t_ber_channel(text):
        return "TC BER"
    if "DDMI" in upper:
        return "DDMI"
    if "3T" in upper and "BER" in upper:
        return "TC BER"
    if "TC" in upper and "BER" in upper:
        return "TC BER"
    if "ATS" in upper:
        return "ATS"
    if "BURN" in upper:
        return "Burn In"
    if "SWITCH" in upper:
        return "Switch"
    if "TP2TP3_LT" in upper or "_LT" in upper or upper.endswith("LT") or upper.startswith("LT"):
        return "LT"
    if "TP2TP3_HT" in upper or "_HT" in upper or upper.endswith("HT") or upper.startswith("HT"):
        return "HT"
    if "TP2TP3_RT" in upper or "_RT" in upper or upper.endswith("RT") or upper.startswith("RT"):
        return "RT"
    return None


def determine_station(row: pd.Series, station_column: str | None) -> str | None:
    if station_column:
        value = row.get(station_column)
        station = normalize_station(str(value)) if value is not None else None
        if station:
            return station
    value = row.get(CH_NUMBER_HEADER)
    return normalize_station(str(value)) if value is not None else None


def is_pass(value: object) -> bool:
    if value is None:
        return False
    text = str(value).strip().upper()
    return text in {"PASS", "P", "OK", "TRUE", "1", "Y", "YES"}


def determine_sort_columns(df: pd.DataFrame) -> list[str]:
    candidates = [
        "test_datetime",
        "TEST_DATETIME",
        "TestDateTime",
        "TESTNUMBER",
        "TestNumber",
        "TEST_TIME",
        "TESTDATE",
    ]
    columns = [col for col in candidates if col in df.columns]
    return columns if columns else []


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_three_t_ber_channel(value: object) -> bool:
    text = normalize_text(value).lower()
    return text in THREE_T_BER_ITEM_SET


def is_tc_ber_channel(value: object) -> bool:
    text = normalize_text(value).upper()
    return "TC" in text and "BER" in text


def has_value(value: object) -> bool:
    text = normalize_text(value)
    return text not in {"", "nan", "none", "null"}


def split_into_tests(group_df: pd.DataFrame, expected_count: int, sort_columns: list[str]) -> list[pd.DataFrame]:
    if sort_columns:
        group_df = group_df.sort_values(sort_columns)
    total_rows = len(group_df)
    if total_rows == 0:
        return []
    tests = []
    for start in range(0, total_rows, expected_count):
        tests.append(group_df.iloc[start : start + expected_count])
    return tests


def extract_error_code(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if text.lower() in {"0", "0.0"}:
        return ""
    prefix = text.split(":", 1)[0].strip()
    if not prefix:
        return ""
    return prefix.split(" ", 1)[0].strip()


def apply_error_codes(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        df[ERROR_CODE_HEADER] = ""
        return df
    df = df.copy()
    failure_code_column = find_failure_code_column(list(df.columns))
    if not failure_code_column:
        print("⚠️ 找不到 FailureCodeID 欄位，Error code 將為空白")
        df[ERROR_CODE_HEADER] = ""
        return df
    df[ERROR_CODE_HEADER] = df[failure_code_column].apply(extract_error_code)
    return df


def build_data_analysis_metrics(df: pd.DataFrame) -> dict[str, dict[str, float]]:
    component_column = find_component_column(list(df.columns))
    result_column = find_result_column(list(df.columns))
    station_column = find_station_column(list(df.columns))
    ch_pass_fail_columns = find_ch_pass_fail_columns(list(df.columns))

    df = df.copy()
    df["_station"] = df.apply(lambda row: determine_station(row, station_column), axis=1)
    df = df[df["_station"].isin(STATION_ORDER)]

    if not result_column:
        print("⚠️ 找不到結果欄位，良率將以 0 計算")
    if not ch_pass_fail_columns:
        print("⚠️ 找不到 CH_Pass_Fail 欄位，TC BER 良率將以 0 計算")

    sort_columns = determine_sort_columns(df)
    metrics: dict[str, dict[str, float]] = {}
    for station in STATION_ORDER:
        station_df = df[df["_station"] == station]
        if station == "TC BER":
            expected_count = 32
        elif station in {"DDMI", "RT", "LT", "HT"}:
            expected_count = 8
        else:
            expected_count = 24
        fpy_input = 0
        fpy_output = 0
        retest_input = 0
        retest_output = 0

        for _, group in station_df.groupby(component_column):
            tests = split_into_tests(group, expected_count, sort_columns)
            if not tests:
                continue
            fpy_input += 1
            if station == "TC BER":
                if ch_pass_fail_columns and all(
                    is_pass(value)
                    for column in ch_pass_fail_columns
                    for value in tests[0][column]
                ):
                    fpy_output += 1
            else:
                if result_column and all(is_pass(val) for val in tests[0][result_column]):
                    fpy_output += 1
            if len(tests) > 1:
                retest_input += len(tests) - 1
                if station == "TC BER":
                    if ch_pass_fail_columns:
                        retest_output += sum(
                            1
                            for test_df in tests[1:]
                            if all(
                                is_pass(value)
                                for column in ch_pass_fail_columns
                                for value in test_df[column]
                            )
                        )
                else:
                    if result_column:
                        retest_output += sum(
                            1 for test_df in tests[1:] if all(is_pass(val) for val in test_df[result_column])
                        )

        fpy_rate = fpy_output / fpy_input if fpy_input else 0
        retest_rate = retest_output / retest_input if retest_input else 0
        metrics[station] = {
            "fpy_input": fpy_input,
            "fpy_output": fpy_output,
            "fpy_rate": fpy_rate,
            "retest_input": retest_input,
            "retest_output": retest_output,
            "retest_rate": retest_rate,
        }
    return metrics


def build_failed_devices(df: pd.DataFrame) -> pd.DataFrame:
    component_column = find_component_column(list(df.columns))
    ch_pass_fail_columns = find_ch_pass_fail_columns(list(df.columns))

    if not ch_pass_fail_columns:
        print("⚠️ 找不到 CH_Pass_Fail 欄位，Failed Device sheet 將為空")
        return df.head(0).drop(columns=["_category"], errors="ignore")

    sort_columns = determine_sort_columns(df)
    failed_tests: list[pd.DataFrame] = []

    for category in ["TC_BER"]:
        category_df = df[df["_category"] == category]
        if category_df.empty:
            continue
        expected_count = 32
        for _, group in category_df.groupby(component_column):
            tests = split_into_tests(group, expected_count, sort_columns)
            for test_df in tests:
                if any(
                    not is_pass(value)
                    for column in ch_pass_fail_columns
                    for value in test_df[column]
                ):
                    failed_record = test_df.copy()
                    failed_record["Category"] = category
                    failed_tests.append(failed_record)

    if failed_tests:
        failed_df = pd.concat(failed_tests, ignore_index=True)
    else:
        failed_df = df.head(0)

    if "BER1" in failed_df.columns:
        failed_df = failed_df[failed_df["BER1"].apply(has_value)]

    return failed_df.drop(columns=["_category"], errors="ignore")


def add_equipment_column(df: pd.DataFrame, equipment_map: dict[str, str]) -> pd.DataFrame:
    if df.empty or not equipment_map:
        return df
    test_column = find_testnumber_column(list(df.columns))
    if not test_column:
        print("⚠️ Failed Device sheet 找不到 TESTNUMBER 欄位，Equipment 會留空")
        return df
    normalized_testnumbers = df[test_column].map(normalize_testnumber)
    equipment_series = normalized_testnumbers.map(equipment_map).fillna("")
    insert_at = df.columns.get_loc(test_column) + 1
    df_with_equipment = df.copy()
    df_with_equipment.insert(insert_at, "Equipment", equipment_series)
    return df_with_equipment


def build_failed_component_records(df: pd.DataFrame) -> pd.DataFrame:
    component_column = find_component_column(list(df.columns))
    station_column = find_station_column(list(df.columns))
    result_column = find_result_column(list(df.columns))
    ch_pass_fail_columns = find_ch_pass_fail_columns(list(df.columns))

    if not ch_pass_fail_columns and not result_column:
        print("⚠️ 找不到結果欄位或 CH_Pass_Fail 欄位，Pareto 統計將為空")
        return pd.DataFrame(columns=["category", "component_id", "error_code"])

    sort_columns = determine_sort_columns(df)
    records: list[dict[str, str]] = []

    df = df.copy()
    df["_station"] = df.apply(lambda row: determine_station(row, station_column), axis=1)
    df = df[df["_station"].isin(STATION_ORDER)]

    for station in STATION_ORDER:
        station_df = df[df["_station"] == station]
        if station_df.empty:
            continue
        if station == "TC BER":
            expected_count = 32
        elif station in {"DDMI", "RT", "LT", "HT"}:
            expected_count = 8
        else:
            expected_count = 24
        for component_id, group in station_df.groupby(component_column):
            tests = split_into_tests(group, expected_count, sort_columns)
            for test_df in tests:
                if station == "TC BER":
                    if not ch_pass_fail_columns:
                        continue
                    failed = any(
                        not is_pass(value)
                        for column in ch_pass_fail_columns
                        for value in test_df[column]
                    )
                else:
                    if not result_column:
                        continue
                    failed = any(not is_pass(val) for val in test_df[result_column])
                if not failed:
                    continue
                error_code = ""
                if ERROR_CODE_HEADER in test_df.columns:
                    for value in test_df[ERROR_CODE_HEADER]:
                        error_code = normalize_text(value)
                        if error_code:
                            break
                records.append(
                    {
                        "category": station,
                        "component_id": str(component_id),
                        "error_code": error_code,
                    }
                )

    return pd.DataFrame(records, columns=["category", "component_id", "error_code"])


def load_output_workbook(base_dir: str) -> Workbook:
    template_path = os.path.join(base_dir, FUNCTION_TEMPLATE)
    if os.path.exists(template_path):
        return load_workbook(template_path)
    workbook = Workbook()
    if workbook.active:
        workbook.remove(workbook.active)
    workbook.create_sheet(DATA_ANALYSIS_SHEET)
    return workbook


def _chunked(values: list[str], size: int) -> list[list[str]]:
    return [values[i : i + size] for i in range(0, len(values), size)]


def build_master_query(sql_text: str, placeholders: str) -> str:
    return f"""
SELECT *
FROM ({sql_text}) AS master
WHERE TESTNUMBER IN ({placeholders});
""".strip()


def fetch_master_equipment_map(
    conn: pyodbc.Connection,
    sql_text: str,
    testnumbers: list[str],
) -> dict[str, str]:
    if not testnumbers:
        return {}
    cleaned_testnumbers = [
        normalized
        for value in testnumbers
        if (normalized := normalize_testnumber(value))
    ]
    if not cleaned_testnumbers:
        return {}
    equipment_map: dict[str, str] = {}
    for chunk in _chunked(cleaned_testnumbers, 900):
        placeholders = ",".join("?" for _ in chunk)
        query = build_master_query(sql_text, placeholders)
        master_df = pd.read_sql_query(query, conn, params=chunk)
        if master_df.empty:
            continue
        test_column = find_testnumber_column(list(master_df.columns))
        equipment_column = find_equipment_column(list(master_df.columns))
        if not test_column or not equipment_column:
            print("⚠️ MASTER.sql 查不到 TESTNUMBER 或 EQUPMENT 欄位，Equipment 會留空")
            return equipment_map
        for _, row in master_df[[test_column, equipment_column]].iterrows():
            test_value = normalize_testnumber(row[test_column])
            if not test_value:
                continue
            equipment_value = row[equipment_column]
            equipment_map[str(test_value)] = "" if pd.isna(equipment_value) else str(equipment_value)
    return equipment_map


def _find_last_data_row(ws) -> int:
    for row_idx in range(ws.max_row, 0, -1):
        if any(cell.value is not None for cell in ws[row_idx]):
            return row_idx
    return 0


def normalize_excel_value(value: object) -> object:
    if isinstance(value, (tuple, list, set, dict)):
        return str(value)
    return value


def write_dataframe_to_sheet(workbook: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
    else:
        ws = workbook.create_sheet(title=sheet_name)
    last_row = _find_last_data_row(ws)
    include_header = last_row == 0
    safe_df = df.applymap(normalize_excel_value)
    for row in dataframe_to_rows(safe_df, index=False, header=include_header):
        ws.append(row)


def build_pareto_table_from_codes(
    codes: pd.Series,
    input_total: float,
    limit: int = PARETO_LIMIT,
) -> pd.DataFrame:
    if codes.empty:
        return pd.DataFrame(columns=PARETO_COLUMNS)

    cleaned = (
        codes
        .map(extract_error_code)
        .map(normalize_text)
        .replace("", pd.NA)
        .dropna()
    )
    if cleaned.empty:
        return pd.DataFrame(columns=PARETO_COLUMNS)

    counts = cleaned.value_counts().head(limit).reset_index()
    if "index" in counts.columns:
        counts = counts.rename(columns={"index": "Error Code", "error_code": "Fail Q'ty"})
    else:
        value_col = "error_code" if "error_code" in counts.columns else counts.columns[0]
        count_col = "count" if "count" in counts.columns else counts.columns[1]
        counts = counts.rename(columns={value_col: "Error Code", count_col: "Fail Q'ty"})
    counts["Fail Q'ty"] = pd.to_numeric(counts["Fail Q'ty"], errors="coerce").fillna(0)
    input_total_value = pd.to_numeric(input_total, errors="coerce")
    if pd.isna(input_total_value):
        input_total_value = 0
    counts["Failed Rate"] = counts["Fail Q'ty"] / input_total_value if input_total_value else 0
    total_fail = counts["Fail Q'ty"].sum()
    counts["Cum%"] = counts["Fail Q'ty"].cumsum() / total_fail if total_fail else 0
    return counts[PARETO_COLUMNS]


def build_pareto_table(
    failed_components: pd.DataFrame,
    station: str,
    input_total: float,
) -> pd.DataFrame:
    if failed_components.empty:
        return pd.DataFrame(columns=PARETO_COLUMNS)

    station_df = failed_components[
        (failed_components["category"] == station)
        & (failed_components["error_code"].str.strip() != "")
    ]
    if station_df.empty:
        return pd.DataFrame(columns=PARETO_COLUMNS)

    return build_pareto_table_from_codes(station_df["error_code"], input_total)


def build_failed_device_pareto_table(
    failed_devices: pd.DataFrame,
    input_total: float,
) -> pd.DataFrame:
    if failed_devices.empty:
        return pd.DataFrame(columns=PARETO_COLUMNS)
    failure_code_column = find_failure_code_column(list(failed_devices.columns))
    if not failure_code_column:
        print("⚠️ 找不到 FailureCodeID 欄位，TC BER Pareto Chart 將為空")
        return pd.DataFrame(columns=PARETO_COLUMNS)
    return build_pareto_table_from_codes(failed_devices[failure_code_column], input_total)


def write_pareto_table(
    ws,
    start_row: int,
    table: pd.DataFrame,
    clear_until_row: int,
    write_rows: bool = True,
) -> None:
    for row in range(start_row + 1, clear_until_row + 1):
        for col in range(1, len(PARETO_COLUMNS) + 1):
            ws.cell(row=row, column=col, value=None)

    if not write_rows:
        return

    for idx, row in enumerate(table.itertuples(index=False), start=start_row + 1):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=idx, column=col_index, value=value)


def populate_data_analysis_sheet(
    workbook: Workbook,
    metrics: dict[str, dict[str, float]],
    failed_devices: pd.DataFrame,
    failed_components: pd.DataFrame,
) -> None:
    if DATA_ANALYSIS_SHEET not in workbook.sheetnames:
        workbook.create_sheet(DATA_ANALYSIS_SHEET)
    ws = workbook[DATA_ANALYSIS_SHEET]

    fpy_row_map = {
        "DDMI": 3,
        "RT": 4,
        "LT": 5,
        "HT": 6,
        "Burn In": 7,
        "TC BER": 9,
        "ATS": 10,
        "Switch": 11,
    }
    retest_row_map = {
        "DDMI": 16,
        "RT": 17,
        "LT": 18,
        "HT": 19,
        "Burn In": 20,
        "TC BER": 22,
        "ATS": 23,
        "Switch": 24,
    }

    for station, row in fpy_row_map.items():
        data = metrics.get(station, {})
        ws[f"B{row}"] = data.get("fpy_input", 0)
        ws[f"C{row}"] = data.get("fpy_output", 0)
        ws[f"D{row}"] = data.get("fpy_rate", 0)

    for station, row in retest_row_map.items():
        data = metrics.get(station, {})
        ws[f"B{row}"] = data.get("retest_input", 0)
        ws[f"C{row}"] = data.get("retest_output", 0)
        ws[f"D{row}"] = data.get("retest_rate", 0)

    pareto_configs = [
        ("DDMI", 28, 41, "components", True),
        ("RT", 43, 56, "components", True),
        ("LT", 58, 71, "components", True),
        ("HT", 73, 86, "components", True),
        ("ATS", 88, 100, "components", True),
        ("TC BER", 103, 115, "failed_devices", False),
        ("TC BER", 118, 130, "components", True),
    ]

    empty_pareto = pd.DataFrame(columns=PARETO_COLUMNS)
    for station, start_row, clear_until_row, source, write_rows in pareto_configs:
        if not write_rows:
            write_pareto_table(
                ws,
                start_row=start_row,
                table=empty_pareto,
                clear_until_row=clear_until_row,
                write_rows=False,
            )
            continue
        input_total = metrics.get(station, {}).get("fpy_input", 0)
        if source == "failed_devices":
            pareto_table = build_failed_device_pareto_table(failed_devices, input_total)
        else:
            pareto_table = build_pareto_table(failed_components, station, input_total)
        write_pareto_table(
            ws,
            start_row=start_row,
            table=pareto_table,
            clear_until_row=clear_until_row,
        )


def main():
    args = parse_args()
    start_date_obj = parse_date(args.start_date)
    end_date_obj = parse_date(args.end_date)
    start_date = start_date_obj.isoformat()
    end_date = end_date_obj.isoformat()
    test_login()

    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = normalize_output_dir(base_dir, args.output_dir)
    os.makedirs(output_dir, exist_ok=True)
    out_path = build_output_path(output_dir, args.start_date, args.end_date)
    out_path = ensure_unique_output_path(out_path)

    try:
        print(f"🚀 連線 DB：{DATABASE}")
        with pyodbc.connect(conn_str_with_db(DATABASE), timeout=30) as conn:

            # 1) 先抓 TOP 0 取得欄位（確認你已「進入報表/view」）
            cur = conn.cursor()
            cur.execute(f"SELECT TOP 0 * FROM {TARGET_OBJECT};")
            cols = [d[0] for d in cur.description]
            print(f"✅ 欄位數：{len(cols)}（已連到該報表 view）")

            # 2) 匯出資料（先小量）
            export_sql = build_export_query(EXPORT_N, base_dir)
            print(f"\n📤 匯出 TOP {EXPORT_N} 到 Excel：{out_path}")
            t0 = time.time()
            df = pd.read_sql_query(
                export_sql,
                conn,
                params=[start_date, end_date],
            )
            df = apply_header(df)
            df = filter_df_by_date_range(df, start_date_obj, end_date_obj)
            print(f"✅ export rows={len(df)} time={time.time()-t0:.1f}s")
            if PREVIEW_N > 0:
                print(f"\n👀 預覽資料 TOP {PREVIEW_N}：")
                with pd.option_context("display.max_columns", 20, "display.width", 180):
                    print(df.head(min(PREVIEW_N, 5)))

            if CH_NUMBER_HEADER not in df.columns:
                raise KeyError(f"查無欄位 {CH_NUMBER_HEADER}")

            equipment_map: dict[str, str] = {}
            testnumber_column = find_testnumber_column(list(df.columns))
            if not testnumber_column:
                print("⚠️ BER 資料查不到 TESTNUMBER 欄位，Equipment 會留空")
            else:
                master_sql_path = os.path.join(base_dir, MASTER_SQL_FILE)
                if os.path.exists(master_sql_path):
                    master_sql = load_sql(master_sql_path)
                    testnumbers = (
                        df[testnumber_column]
                        .dropna()
                        .astype(str)
                        .unique()
                        .tolist()
                    )
                    equipment_map = fetch_master_equipment_map(conn, master_sql, testnumbers)
                else:
                    print("⚠️ 找不到 MASTER.sql，Equipment 會留空")

            categories = ["TC_BER", "其他"]
            df["_category"] = df[CH_NUMBER_HEADER].apply(classify_ch_number)
            workbook = load_output_workbook(base_dir)
            df = apply_error_codes(df)
            analysis_df = df
            metrics = build_data_analysis_metrics(analysis_df)
            failed_devices = build_failed_devices(analysis_df)
            failed_devices = add_equipment_column(failed_devices, equipment_map)
            failed_components = build_failed_component_records(analysis_df)
            for category in categories:
                sheet_df = df[df["_category"] == category].drop(columns=["_category"])
                if sheet_df.empty:
                    sheet_df = df.head(0).drop(columns=["_category"])
                if category == "TC_BER" and equipment_map:
                    test_column = find_testnumber_column(list(sheet_df.columns))
                    if test_column:
                        sheet_df = sheet_df.copy()
                        normalized_testnumbers = sheet_df[test_column].map(normalize_testnumber)
                        sheet_df["Equipment"] = (
                            normalized_testnumbers
                            .map(equipment_map)
                            .fillna("")
                        )
                    else:
                        print("⚠️ TC_BER sheet 找不到 TESTNUMBER 欄位，Equipment 會留空")
                write_dataframe_to_sheet(workbook, category, sheet_df)

            write_dataframe_to_sheet(workbook, "Failed Device", failed_devices)
            populate_data_analysis_sheet(workbook, metrics, failed_devices, failed_components)
            workbook.save(out_path)

            print("📁 Excel 已輸出：", out_path)

    except Exception as e:
        print("❌ 查詢或匯出失敗：")
        print(e)
        sys.exit(2)


if __name__ == "__main__":
    main()
