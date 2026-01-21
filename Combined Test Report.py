# -*- coding: utf-8 -*-

import argparse
import datetime
import importlib.util
import os
import sys
import time

import pandas as pd
import pyodbc
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

OUTPUT_EXTENSION = ".xlsx"
FUNCTION_TEMPLATE = "Function.xlsx"
DATA_ANALYSIS_SHEET = "Data Analysis"
EQUIPMENT_STATUS_SHEET = "equiment status"
ERROR_CODE_HEADER_DEFAULT = "Error code"
STATION_NAME_HEADER = "STATION NAME"
ERROR_CODE_CANONICAL_HEADER = "Error Code"
EQUIPMENT_PERFORMANCE_SHEET = "Equipment Performance"
EQUIPMENT_PERFORMANCE_DATA_SHEET = "Equipment Performance Data"
EQUIPMENT_PERFORMANCE_DDMI_ITEMS = [
    "Power(dBm)",
    "Rxp_Slope",
    "Txp_Slope",
    "Vcc_Slope",
]
EQUIPMENT_PERFORMANCE_ATS_ITEMS = [
    "dTxP",
    "dRxP1",
    "dVcc(%)",
]
EQUIPMENT_PERFORMANCE_TH_ITEMS = [
    "dTxP",
    "dRxP1",
    "dVcc(%)",
]
STATION_NAME_MAP = {
    "ATS": {
        "T157100002205_1": "ATS1_L",
        "T157100002205_2": "ATS1_R",
        "T157100002022_1": "ATS2_L",
        "T157100002022_2": "ATS2_R",
        "T157100002072_1": "ATS3_L",
        "T157100002072_2": "ATS3_R",
        "T157100002201_1": "ATS4_L",
        "T157100002201_2": "ATS4_R",
        "T157100002402_1": "ATS5_L",
        "T157100002402_2": "ATS5_R",
        "T157100002535_1": "ATS6_L",
        "T157100002535_2": "ATS6_R",
        "T157100002533_1": "ATS7_L",
        "T157100002533_2": "ATS7_R",
        "T157100002329_1": "OQC_L",
        "T157100002329_2": "OQC_R",
    },
    "RT": {
        "T157100002534": "RT2",
        "T157100002113": "RT1",
    },
    "DDMI": {
        "T157100002101_1": "DDMI 1_L",
        "T157100002101_2": "DDMI 1_R",
        "T157100002316_1": "DDMI 2_L",
        "T157100002316_2": "DDMI 2_R",
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--start-date", required=True, help="YYYY-MM-DD")
    parser.add_argument("--end-date", required=True, help="YYYY-MM-DD")
    parser.add_argument("--output-path", help="輸出檔案完整路徑 (可省略副檔名)")
    return parser.parse_args()


def parse_date(value: str) -> datetime.date:
    try:
        return datetime.date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError("日期格式需為 YYYY-MM-DD") from exc


def format_date_range(start_date: str, end_date: str) -> str:
    if start_date == end_date:
        return start_date
    return f"{start_date}_{end_date}"


def build_output_path(base_dir: str, start_date: str, end_date: str) -> str:
    base_name = os.path.splitext(os.path.basename(__file__))[0]
    date_range = format_date_range(start_date, end_date)
    filename = f"{base_name}_{date_range}{OUTPUT_EXTENSION}"
    return os.path.join(base_dir, filename)


def normalize_output_path(
    base_dir: str,
    output_path: str | None,
    start_date: str,
    end_date: str,
) -> str:
    if not output_path:
        return build_output_path(base_dir, start_date, end_date)

    expanded_path = os.path.expanduser(output_path)
    if not os.path.isabs(expanded_path):
        expanded_path = os.path.join(base_dir, expanded_path)

    if os.path.isdir(expanded_path):
        filename = os.path.basename(build_output_path(base_dir, start_date, end_date))
        return os.path.join(expanded_path, filename)

    root, ext = os.path.splitext(expanded_path)
    if ext.lower() != OUTPUT_EXTENSION:
        return f"{expanded_path}{OUTPUT_EXTENSION}"
    return expanded_path


def ensure_unique_output_path(path: str) -> str:
    if not os.path.exists(path):
        return path
    root, ext = os.path.splitext(path)
    counter = 1
    while True:
        candidate = f"{root}_{counter}{ext}"
        if not os.path.exists(candidate):
            return candidate
        counter += 1


def load_module(module_path: str, module_name: str):
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    if spec is None or spec.loader is None:
        raise ImportError(f"無法載入模組：{module_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def load_template_workbook(base_dir: str) -> Workbook:
    template_path = os.path.join(base_dir, FUNCTION_TEMPLATE)
    if os.path.exists(template_path):
        return load_workbook(template_path)
    workbook = Workbook()
    if workbook.active:
        workbook.remove(workbook.active)
    workbook.create_sheet(DATA_ANALYSIS_SHEET)
    return workbook


def ensure_data_analysis_template(workbook: Workbook) -> object:
    if DATA_ANALYSIS_SHEET in workbook.sheetnames:
        return workbook[DATA_ANALYSIS_SHEET]
    return workbook.create_sheet(DATA_ANALYSIS_SHEET)


def set_cell_value(ws, row: int, column: int, value: object) -> None:
    cell = ws.cell(row=row, column=column)
    if not isinstance(cell, MergedCell):
        cell.value = value
        return
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            if row == merged_range.min_row and column == merged_range.min_col:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col, value=value)
            return


def export_report_dataframe(module, start_date: str, end_date: str) -> pd.DataFrame:
    with pyodbc.connect(module.conn_str_with_db(module.DATABASE), timeout=30) as conn:
        if hasattr(module, "build_columns_query"):
            use_openquery = False
            cur = conn.cursor()
            try:
                cur.execute(module.build_columns_query(False))
            except pyodbc.Error:
                use_openquery = True
                cur.execute(module.build_columns_query(True))

            if use_openquery and hasattr(module, "build_sorted_query_openquery"):
                export_sql = module.build_sorted_query_openquery(module.EXPORT_N, start_date, end_date)
                df = pd.read_sql_query(export_sql, conn)
            else:
                export_sql = module.build_sorted_query(module.EXPORT_N)
                df = pd.read_sql_query(export_sql, conn, params=[start_date, end_date])
        else:
            export_sql = module.build_sorted_query(module.EXPORT_N)
            df = pd.read_sql_query(export_sql, conn, params=[start_date, end_date])

    return module.apply_header(df)


def find_location_column(columns: list[str]) -> str | None:
    candidates = [
        "location",
        "testlocation",
        "test_location",
        "stationlocation",
        "station_location",
        "site",
        "testsite",
        "test_site",
        "line",
        "testline",
        "test_line",
    ]
    return next((col for col in columns if col.lower() in candidates), None)


def find_equipment_column(columns: list[str]) -> str | None:
    return next((col for col in columns if col.strip().lower() == "equipment"), None)


def _normalize_column_key(value: str) -> str:
    return "".join(ch for ch in value.lower() if ch.isalnum())


def find_station_name_column(columns: list[str]) -> str | None:
    target_key = _normalize_column_key(STATION_NAME_HEADER)
    for column in columns:
        if _normalize_column_key(str(column)) == target_key:
            return column
    candidates = [
        "station",
        "teststation",
        "test_station",
        "station_name",
        "station name",
        "stationname",
    ]
    for column in columns:
        if str(column).strip().lower() in candidates:
            return column
    return None


def resolve_station_name(category: object, equipment: object) -> str:
    category_key = str(category).strip()
    equipment_key = str(equipment).strip()
    if not category_key or not equipment_key:
        return ""
    return STATION_NAME_MAP.get(category_key, {}).get(equipment_key, "")


def build_station_name_series(
    df: pd.DataFrame,
    category_column: str,
    equipment_column: str,
) -> pd.Series:
    station_column = find_station_name_column(list(df.columns))
    if station_column and station_column in df.columns:
        station_series = df[station_column].astype(str).str.strip()
    else:
        station_series = df.apply(
            lambda row: resolve_station_name(row.get(category_column, ""), row.get(equipment_column, "")),
            axis=1,
        )
    fallback = df[category_column].astype(str).str.strip()
    station_series = station_series.fillna("").astype(str).str.strip()
    return station_series.mask(station_series == "", fallback)


def _find_equipment_performance_column(columns: list[str], target: str) -> str | None:
    target_key = _normalize_column_key(target)
    for column in columns:
        if _normalize_column_key(str(column)) == target_key:
            return column
    return None


def _build_station_name_order(series: pd.Series) -> list[str]:
    station_names: list[str] = []
    for value in series.dropna():
        name = str(value).strip()
        if not name or name in station_names:
            continue
        station_names.append(name)
    return station_names


def _build_location_order(series: pd.Series) -> list[str]:
    location_names: list[str] = []
    for value in series.fillna(""):
        name = str(value).strip()
        if not name or name in location_names:
            continue
        location_names.append(name)
    return location_names


def resolve_equipment_warning(sheet_prefix: str) -> str:
    if sheet_prefix == "800G_TRX":
        return "⚠️ TRX 資料查不到 TESTNUMBER 欄位，Equipment 會留空"
    if sheet_prefix == "800G_Fixed_BER":
        return "⚠️ 3T BER 資料查不到 TESTNUMBER 欄位，Equipment 會留空"
    if sheet_prefix == "BER_Symbol_Error":
        return "⚠️ BER 資料查不到 TESTNUMBER 欄位，Equipment 會留空"
    return "⚠️ 查不到 TESTNUMBER 欄位，Equipment 會留空"


def fetch_equipment_map(module, df: pd.DataFrame, base_dir: str, sheet_prefix: str) -> dict[str, str]:
    if not hasattr(module, "fetch_master_equipment_map"):
        return {}
    testnumber_finder = getattr(module, "find_testnumber_column", None)
    if not callable(testnumber_finder):
        return {}
    testnumber_column = testnumber_finder(list(df.columns))
    if not testnumber_column:
        print(resolve_equipment_warning(sheet_prefix))
        return {}
    master_sql_file = getattr(module, "MASTER_SQL_FILE", "MASTER.sql")
    master_sql_path = os.path.join(base_dir, master_sql_file)
    if not os.path.exists(master_sql_path):
        print("⚠️ 找不到 MASTER.sql，Equipment 會留空")
        return {}
    load_sql = getattr(module, "load_sql", None)
    if callable(load_sql):
        master_sql = load_sql(master_sql_path)
    else:
        with open(master_sql_path, "r", encoding="utf-8") as file:
            master_sql = file.read()
    testnumbers = (
        df[testnumber_column]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )
    if not testnumbers:
        return {}
    with pyodbc.connect(module.conn_str_with_db(module.DATABASE), timeout=30) as conn:
        return module.fetch_master_equipment_map(conn, master_sql, testnumbers)


def should_add_equipment(sheet_prefix: str, category: str) -> bool:
    if sheet_prefix == "800G_TRX":
        return category != "其他"
    if sheet_prefix == "800G_Fixed_BER":
        return category == "3T_BER"
    if sheet_prefix == "BER_Symbol_Error":
        return category == "TC_BER"
    return False


def normalize_excel_value(value: object) -> object:
    if isinstance(value, (tuple, list, set, dict)):
        return str(value)
    return value


def write_dataframe_to_sheet(workbook: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    ws = workbook.create_sheet(title=sheet_name)
    safe_df = df.applymap(normalize_excel_value)
    for row in dataframe_to_rows(safe_df, index=False, header=True):
        ws.append(row)


def _measure_cell_length(value: object) -> int:
    if value is None:
        return 0
    text = str(value)
    return max((len(line) for line in text.splitlines()), default=0)


def format_equipment_status_sheet(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="BDD7EE")
    header_attention_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    yield_rate_header_fill = PatternFill(fill_type="solid", fgColor="FF0000")
    yield_green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    yield_orange_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    yield_red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    red_border = Border(
        left=Side(style="thin", color="FF0000"),
        right=Side(style="thin", color="FF0000"),
        top=Side(style="thin", color="FF0000"),
        bottom=Side(style="thin", color="FF0000"),
    )
    base_font = Font(name="Calibri", size=10)
    header_font = Font(name="Calibri", size=10, bold=True)
    header_attention_font = Font(name="Calibri", size=10, bold=True, color="9C0006")
    yield_rate_header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.font = base_font
            cell.border = border

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    yield_rate_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Yield rate":
            yield_rate_col = idx
            break

    if yield_rate_col:
        header_cell = ws.cell(row=1, column=yield_rate_col)
        header_cell.fill = yield_rate_header_fill
        header_cell.font = yield_rate_header_font
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=yield_rate_col)
            cell.number_format = "0.0%"
            if isinstance(cell.value, (int, float)):
                if cell.value >= 0.9:
                    cell.fill = yield_green_fill
                elif cell.value >= 0.7:
                    cell.fill = yield_orange_fill
                else:
                    cell.fill = yield_red_fill

    known_columns = {
        "Report",
        "Category",
        "Equipment",
        STATION_NAME_HEADER,
        "Location",
        "fpy_input",
        "fpy_output",
        "Yield rate",
    }
    error_code_columns = [
        (idx, cell.value)
        for idx, cell in enumerate(ws[1], start=1)
        if cell.value not in known_columns
    ]
    if error_code_columns:
        for row_idx in range(2, ws.max_row + 1):
            numeric_values = []
            for column_idx, _header in error_code_columns:
                value = ws.cell(row=row_idx, column=column_idx).value
                if isinstance(value, (int, float)):
                    numeric_values.append((column_idx, value))
            if not numeric_values:
                continue
            max_value = max(value for _column_idx, value in numeric_values)
            for column_idx, value in numeric_values:
                if value == max_value:
                    ws.cell(row=row_idx, column=column_idx).border = red_border

    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        max_length = max((_measure_cell_length(cell.value) for cell in column_cells), default=0)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 8), 60)


def reorder_error_code_column(
    df: pd.DataFrame,
    error_code_header: str,
    insert_after_index: int = 8,
) -> pd.DataFrame:
    if error_code_header not in df.columns:
        return df
    columns = list(df.columns)
    if columns.index(error_code_header) == insert_after_index:
        return df
    columns.remove(error_code_header)
    insert_position = min(insert_after_index, len(columns))
    columns.insert(insert_position, error_code_header)
    return df[columns]


def move_sheet_after(workbook: Workbook, sheet_name: str, after_sheet_name: str) -> None:
    if sheet_name not in workbook.sheetnames or after_sheet_name not in workbook.sheetnames:
        return
    sheet = workbook[sheet_name]
    after_sheet = workbook[after_sheet_name]
    workbook._sheets.remove(sheet)
    insert_index = workbook._sheets.index(after_sheet) + 1
    workbook._sheets.insert(insert_index, sheet)


def hide_sheet(workbook: Workbook, sheet_name: str) -> None:
    if sheet_name in workbook.sheetnames:
        workbook[sheet_name].sheet_state = "hidden"


def _add_equipment_column(df: pd.DataFrame, module, equipment_map: dict[str, str]) -> pd.DataFrame:
    if equipment_map and hasattr(module, "add_equipment_column"):
        return module.add_equipment_column(df, equipment_map)
    if "Equipment" not in df.columns:
        df = df.copy()
        df["Equipment"] = ""
    return df


def _add_location_column(df: pd.DataFrame, location_column: str | None) -> pd.DataFrame:
    df = df.copy()
    if location_column and location_column in df.columns:
        df["Location"] = df[location_column].fillna("")
    else:
        df["Location"] = ""
    return df


def populate_equipment_performance_section(
    data_ws,
    chart_ws,
    df: pd.DataFrame,
    categories: list[str],
    items: list[str],
    data_start_row: int,
    chart_start_row: int,
    section_title: str,
    group_with_location: bool = False,
    split_by_location: bool = False,
    component_column: str | None = None,
    sort_columns: list[str] | None = None,
) -> tuple[int, int]:
    section_df = df[df["_category"].isin(categories)].copy()
    if section_df.empty:
        print(f"⚠️ {section_title} 無資料，Equipment Performance 將跳過")
        return data_start_row, chart_start_row

    equipment_column = find_equipment_column(list(section_df.columns))
    if not equipment_column:
        print(f"⚠️ {section_title} 找不到 Equipment 欄位，Equipment Performance 將跳過")
        return data_start_row, chart_start_row

    if group_with_location and "Location" not in section_df.columns:
        section_df["Location"] = ""

    if component_column not in section_df.columns:
        component_column = None

    section_df[STATION_NAME_HEADER] = build_station_name_series(
        section_df,
        category_column="_category",
        equipment_column=equipment_column,
    )
    station_names = _build_station_name_order(section_df[STATION_NAME_HEADER])
    if not station_names:
        print(f"⚠️ {section_title} Station Name 無有效資料，Equipment Performance 將跳過")
        return data_start_row, chart_start_row

    rename_map: dict[str, str] = {}
    for item in items:
        matched = _find_equipment_performance_column(list(section_df.columns), item)
        if matched:
            rename_map[matched] = item
        else:
            section_df[item] = None
    if rename_map:
        section_df = section_df.rename(columns=rename_map)

    for item in items:
        section_df[item] = pd.to_numeric(section_df[item], errors="coerce")

    sort_columns = [col for col in (sort_columns or []) if col in section_df.columns]

    data_ws.cell(row=data_start_row, column=1, value=section_title)
    chart_ws.cell(row=chart_start_row, column=1, value=section_title)
    data_block_start_row = data_start_row + 1
    chart_block_start_row = chart_start_row + 1
    chart_columns = [8, 16, 24, 32]
    chart_height_rows = 15

    def configure_chart_axes(chart: LineChart, x_title: str, y_title: str) -> None:
        chart.x_axis.title = x_title
        chart.y_axis.title = y_title
        chart.x_axis.tickLblPos = "low"
        chart.y_axis.tickLblPos = "nextTo"
        chart.x_axis.majorTickMark = "out"
        chart.y_axis.majorTickMark = "out"
        chart.x_axis.minorTickMark = "none"
        chart.y_axis.minorTickMark = "none"
        chart.x_axis.delete = False
        chart.y_axis.delete = False

    for station in station_names:
        station_summary = section_df[section_df[STATION_NAME_HEADER] == station].copy()
        if group_with_location:
            station_summary["Location"] = station_summary["Location"].fillna("").astype(str).str.strip()

        if group_with_location and split_by_location:
            location_names = _build_location_order(station_summary["Location"])
            if not location_names:
                location_names = [""]
        else:
            location_names = [None]

        for location in location_names:
            if location is None:
                location_summary = station_summary
                title_prefix = station
            else:
                location_summary = station_summary[station_summary["Location"] == location].copy()
                title_prefix = f"{station} {location}".strip() if location else station

            if group_with_location and not split_by_location:
                location_summary["Equipment Display"] = location_summary.apply(
                    lambda row: " ".join(
                        part
                        for part in [str(row.get(equipment_column, "")).strip(), str(row.get("Location", "")).strip()]
                        if part
                    ),
                    axis=1,
                )
                display_column = "Equipment Display"
                header_label = "Equipment / Location"
            else:
                display_column = equipment_column
                header_label = "Equipment"

            if component_column:
                display_column = component_column
                header_label = "COMPONENTID"

            sort_fields = [display_column, *sort_columns]
            location_summary = location_summary.sort_values(sort_fields) if sort_fields else location_summary
            if location_summary.empty:
                location_summary = pd.DataFrame(
                    [{display_column: "", **{item: None for item in items}}]
                )

            data_columns = [display_column, *items]
            location_summary = location_summary[data_columns]

            title_row = data_block_start_row
            header_row = data_block_start_row + 1
            data_row_start = data_block_start_row + 2

            data_ws.cell(row=title_row, column=1, value=title_prefix)
            data_ws.cell(row=header_row, column=1, value=header_label)
            for idx, item in enumerate(items, start=2):
                data_ws.cell(row=header_row, column=idx, value=item)

            for row_offset, row in enumerate(location_summary.itertuples(index=False), start=0):
                for col_offset, value in enumerate(row, start=1):
                    data_ws.cell(row=data_row_start + row_offset, column=col_offset, value=value)

            data_row_end = data_row_start + len(location_summary) - 1
            categories_ref = Reference(data_ws, min_col=1, min_row=data_row_start, max_row=data_row_end)

            chart_col_index = 0
            for item_index, item in enumerate(items):
                if not location_summary[item].notna().any():
                    continue
                if chart_col_index >= len(chart_columns):
                    break
                data_ref = Reference(
                    data_ws,
                    min_col=item_index + 2,
                    min_row=header_row,
                    max_row=data_row_end,
                )
                chart = LineChart()
                chart.title = f"{title_prefix} {item}"
                configure_chart_axes(chart, header_label, item)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(categories_ref)
                chart.height = 7
                chart.width = 14
                anchor_col = get_column_letter(chart_columns[chart_col_index])
                chart_ws.add_chart(chart, f"{anchor_col}{chart_block_start_row}")
                chart_col_index += 1

            table_height = len(location_summary) + 2
            data_block_start_row += table_height + 3
            chart_block_start_row += (chart_height_rows if chart_col_index else table_height) + 3

    return data_block_start_row + 1, chart_block_start_row + 1


def _compute_group_fpy(
    df: pd.DataFrame,
    group_fields: list[str],
    module,
    expected_count: int,
    use_ch_pass_fail: bool = False,
) -> list[dict[str, object]]:
    component_column = module.find_component_column(list(df.columns))
    result_column = module.find_result_column(list(df.columns))
    ch_pass_fail_columns = module.find_ch_pass_fail_columns(list(df.columns))
    sort_columns = module.determine_sort_columns(df)

    if use_ch_pass_fail and not ch_pass_fail_columns:
        print("⚠️ 找不到 CH_Pass_Fail 欄位，良率將以 0 計算")
    if not use_ch_pass_fail and not result_column:
        print("⚠️ 找不到結果欄位，良率將以 0 計算")

    rows: list[dict[str, object]] = []
    grouped = df.groupby(group_fields, dropna=False) if group_fields else [((), df)]
    for keys, group_df in grouped:
        fpy_input = 0
        fpy_output = 0
        for _, component_group in group_df.groupby(component_column):
            tests = module.split_into_tests(component_group, expected_count, sort_columns)
            if not tests:
                continue
            fpy_input += 1
            passed = False
            if use_ch_pass_fail:
                if ch_pass_fail_columns:
                    passed = all(
                        module.is_pass(value)
                        for column in ch_pass_fail_columns
                        for value in tests[0][column]
                    )
            else:
                if result_column:
                    passed = all(module.is_pass(val) for val in tests[0][result_column])
            if passed:
                fpy_output += 1
        fpy_rate = fpy_output / fpy_input if fpy_input else 0
        if group_fields:
            if len(group_fields) == 1:
                key_map = {group_fields[0]: keys}
            else:
                key_map = dict(zip(group_fields, keys))
        else:
            key_map = {}
        rows.append(
            {
                **key_map,
                "fpy_input": fpy_input,
                "fpy_output": fpy_output,
                "fpy_rate": fpy_rate,
            }
        )
    return rows


def build_error_code_summary(report_results: dict[str, dict[str, object]]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    report_name_map = {
        "800G_TRX": "800G_TRX_TEST",
        "800G_Fixed_BER": "800G_Fixed_BER_Test",
        "BER_Symbol_Error": "BER_Symbol_Error_Test",
    }

    for sheet_prefix, result in report_results.items():
        failed_devices = result.get("failed_devices")
        if failed_devices is None or failed_devices.empty:
            continue
        module = result["module"]
        equipment_map = result.get("equipment_map", {})
        df = failed_devices.copy()
        if "Equipment" not in df.columns:
            df = _add_equipment_column(df, module, equipment_map)
        location_column = find_location_column(list(df.columns))
        df = _add_location_column(df, location_column)
        error_code_header = getattr(module, "ERROR_CODE_HEADER", ERROR_CODE_HEADER_DEFAULT)
        if error_code_header not in df.columns:
            continue
        df[ERROR_CODE_CANONICAL_HEADER] = df[error_code_header].fillna("").astype(str).str.strip()
        df = df[df[ERROR_CODE_CANONICAL_HEADER] != ""]
        if df.empty:
            continue
        if "Category" not in df.columns:
            df["Category"] = ""
        df["Report"] = report_name_map.get(sheet_prefix, sheet_prefix)
        frames.append(df[["Report", "Category", "Equipment", "Location", ERROR_CODE_CANONICAL_HEADER]])

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True)
    counts = (
        combined.groupby(
            ["Report", "Category", "Equipment", "Location", ERROR_CODE_CANONICAL_HEADER],
            dropna=False,
        )
        .size()
        .reset_index(name="count")
    )
    summary = counts.pivot_table(
        index=["Report", "Category", "Equipment", "Location"],
        columns=ERROR_CODE_CANONICAL_HEADER,
        values="count",
        fill_value=0,
        aggfunc="sum",
    ).reset_index()
    return summary


def build_equipment_status_table(report_results: dict[str, dict[str, object]]) -> pd.DataFrame:
    rows: list[dict[str, object]] = []

    trx_result = report_results.get("800G_TRX")
    if trx_result:
        module = trx_result["module"]
        equipment_map = trx_result.get("equipment_map", {})
        df = trx_result["df"].copy()
        df = _add_equipment_column(df, module, equipment_map)
        location_column = find_location_column(list(df.columns))
        if not location_column:
            print("⚠️ TRX 資料查不到 Location 欄位，Location 會留空")
        df = _add_location_column(df, location_column)
        for category in ["ATS", "DDMI", "RT", "LT", "HT"]:
            category_df = df[df["_category"] == category]
            if category_df.empty:
                continue
            expected_count = 24 if category == "ATS" else 8
            group_fields = ["Equipment", "Location"]
            for row in _compute_group_fpy(category_df, group_fields, module, expected_count):
                row.update(
                    {
                        "Report": "800G_TRX_TEST",
                        "Category": category,
                    }
                )
                if "Location" not in row:
                    row["Location"] = ""
                row[STATION_NAME_HEADER] = resolve_station_name(row.get("Category", ""), row.get("Equipment", ""))
                row["Yield rate"] = row.pop("fpy_rate")
                rows.append(row)

    fixed_result = report_results.get("800G_Fixed_BER")
    if fixed_result:
        module = fixed_result["module"]
        equipment_map = fixed_result.get("equipment_map", {})
        df = fixed_result["df"].copy()
        df = _add_equipment_column(df, module, equipment_map)
        location_column = find_location_column(list(df.columns))
        if not location_column:
            print("⚠️ 3T BER 資料查不到 Location 欄位，Location 會留空")
        df = _add_location_column(df, location_column)
        category_df = df[df["_category"] == "3T_BER"]
        if not category_df.empty:
            for row in _compute_group_fpy(category_df, ["Equipment", "Location"], module, 32, use_ch_pass_fail=True):
                row.update(
                    {
                        "Report": "800G_Fixed_BER_Test",
                        "Category": "3T_BER",
                    }
                )
                row[STATION_NAME_HEADER] = resolve_station_name(row.get("Category", ""), row.get("Equipment", ""))
                row["Yield rate"] = row.pop("fpy_rate")
                rows.append(row)

    symbol_result = report_results.get("BER_Symbol_Error")
    if symbol_result:
        module = symbol_result["module"]
        equipment_map = symbol_result.get("equipment_map", {})
        df = symbol_result["df"].copy()
        df = _add_equipment_column(df, module, equipment_map)
        location_column = find_location_column(list(df.columns))
        if not location_column:
            print("⚠️ BER 資料查不到 Location 欄位，Location 會留空")
        df = _add_location_column(df, location_column)
        category_df = df[df["_category"] == "TC_BER"]
        if not category_df.empty:
            for row in _compute_group_fpy(category_df, ["Equipment", "Location"], module, 32, use_ch_pass_fail=True):
                row.update(
                    {
                        "Report": "BER_Symbol_Error_Test",
                        "Category": "TC_BER",
                    }
                )
                row[STATION_NAME_HEADER] = resolve_station_name(row.get("Category", ""), row.get("Equipment", ""))
                row["Yield rate"] = row.pop("fpy_rate")
                rows.append(row)

    if not rows:
        return pd.DataFrame(
            columns=[
                "Report",
                "Category",
                "Equipment",
                STATION_NAME_HEADER,
                "Location",
                "fpy_input",
                "fpy_output",
                "Yield rate",
            ]
        )

    table = pd.DataFrame(rows)
    error_code_summary = build_error_code_summary(report_results)
    if not error_code_summary.empty:
        table = table.merge(
            error_code_summary,
            on=["Report", "Category", "Equipment", "Location"],
            how="left",
        )

    column_order = [
        "Report",
        "Category",
        "Equipment",
        STATION_NAME_HEADER,
        "Location",
        "fpy_input",
        "fpy_output",
        "Yield rate",
    ]
    for column in column_order:
        if column not in table.columns:
            table[column] = ""
    error_code_columns = [
        column
        for column in table.columns
        if column not in column_order
    ]
    if error_code_columns:
        table[error_code_columns] = table[error_code_columns].fillna(0).astype(int)
    return table[column_order + error_code_columns]


def build_report(
    module,
    sheet_prefix: str,
    categories: list[str],
    category_builder,
    workbook: Workbook,
    start_date: str,
    end_date: str,
    base_dir: str,
) -> dict[str, object]:
    print(f"\n🚀 開始整合：{sheet_prefix}")
    t0 = time.time()

    df = export_report_dataframe(module, start_date, end_date)
    print(f"✅ export rows={len(df)} time={time.time()-t0:.1f}s")

    if module.CH_NUMBER_HEADER not in df.columns:
        raise KeyError(f"查無欄位 {module.CH_NUMBER_HEADER}")

    df["_category"] = category_builder(df, module)
    equipment_map = fetch_equipment_map(module, df, base_dir, sheet_prefix)
    df = module.apply_error_codes(df)
    metrics = module.build_data_analysis_metrics(df)
    failed_devices = module.build_failed_devices(df)
    if equipment_map and hasattr(module, "add_equipment_column"):
        failed_devices = module.add_equipment_column(failed_devices, equipment_map)
    failed_components = module.build_failed_component_records(df)

    for category in categories:
        sheet_df = df[df["_category"] == category].drop(columns=["_category"])
        if sheet_df.empty:
            sheet_df = df.head(0).drop(columns=["_category"])
        if equipment_map and hasattr(module, "add_equipment_column") and should_add_equipment(sheet_prefix, category):
            sheet_df = module.add_equipment_column(sheet_df, equipment_map)
        error_code_header = getattr(module, "ERROR_CODE_HEADER", ERROR_CODE_HEADER_DEFAULT)
        sheet_df = reorder_error_code_column(sheet_df, error_code_header)
        module.write_dataframe_to_sheet(workbook, f"{sheet_prefix} {category}", sheet_df)

    error_code_header = getattr(module, "ERROR_CODE_HEADER", ERROR_CODE_HEADER_DEFAULT)
    failed_devices = reorder_error_code_column(failed_devices, error_code_header)
    module.write_dataframe_to_sheet(workbook, f"{sheet_prefix} Failed Device", failed_devices)
    return {
        "df": df,
        "equipment_map": equipment_map,
        "module": module,
        "metrics": metrics,
        "failed_devices": failed_devices,
        "failed_components": failed_components,
    }


def populate_combined_data_analysis(workbook: Workbook, report_results: dict[str, dict[str, object]]) -> None:
    ws = ensure_data_analysis_template(workbook)

    trx_result = report_results.get("800G_TRX")
    fixed_result = report_results.get("800G_Fixed_BER")
    symbol_result = report_results.get("BER_Symbol_Error")

    combined_metrics: dict[str, dict[str, float]] = {}
    if trx_result:
        combined_metrics.update(trx_result["metrics"])
    if fixed_result:
        combined_metrics["3T BER"] = fixed_result["metrics"].get("3T BER", {})
    if symbol_result:
        combined_metrics["TC BER"] = symbol_result["metrics"].get("TC BER", {})

    fpy_row_map = {
        "DDMI": 3,
        "RT": 4,
        "LT": 5,
        "HT": 6,
        "Burn In": 7,
        "3T BER": 8,
        "TC BER": 9,
        "ATS": 10,
        "Switch": 11,
    }
    retest_row_map = {
        "DDMI": 16,
        "RT": 17,
        "LT": 18,
        "HT": 19,
        "Burn In": 20,
        "3T BER": 21,
        "TC BER": 22,
        "ATS": 23,
        "Switch": 24,
    }

    for station, row in fpy_row_map.items():
        data = combined_metrics.get(station, {})
        set_cell_value(ws, row, 2, data.get("fpy_input", 0))
        set_cell_value(ws, row, 3, data.get("fpy_output", 0))
        set_cell_value(ws, row, 4, data.get("fpy_rate", 0))

    for station, row in retest_row_map.items():
        data = combined_metrics.get(station, {})
        set_cell_value(ws, row, 2, data.get("retest_input", 0))
        set_cell_value(ws, row, 3, data.get("retest_output", 0))
        set_cell_value(ws, row, 4, data.get("retest_rate", 0))

    if trx_result:
        trx_module = trx_result["module"]
        pareto_configs = [
            ("DDMI", 28, 41),
            ("RT", 43, 56),
            ("LT", 58, 71),
            ("HT", 73, 86),
            ("ATS", 88, 100),
        ]
        for station, start_row, clear_until_row in pareto_configs:
            input_total = trx_result["metrics"].get(station, {}).get("fpy_input", 0)
            pareto_table = trx_module.build_pareto_table(
                trx_result["failed_components"],
                station,
                input_total,
            )
            trx_module.write_pareto_table(
                ws,
                start_row=start_row,
                table=pareto_table,
                clear_until_row=clear_until_row,
            )

    if fixed_result:
        fixed_module = fixed_result["module"]
        input_total = fixed_result["metrics"].get("3T BER", {}).get("fpy_input", 0)
        pareto_table = fixed_module.build_failed_device_pareto_table(
            fixed_result["failed_devices"],
            input_total,
        )
        fixed_module.write_pareto_table(
            ws,
            start_row=103,
            table=pareto_table,
            clear_until_row=115,
        )

    if symbol_result:
        symbol_module = symbol_result["module"]
        input_total = symbol_result["metrics"].get("TC BER", {}).get("fpy_input", 0)
        pareto_table = symbol_module.build_pareto_table(
            symbol_result["failed_components"],
            "TC BER",
            input_total,
        )
        symbol_module.write_pareto_table(
            ws,
            start_row=118,
            table=pareto_table,
            clear_until_row=130,
        )


def populate_equipment_performance_sheet(workbook: Workbook, report_results: dict[str, dict[str, object]]) -> None:
    trx_result = report_results.get("800G_TRX")
    if not trx_result:
        print("⚠️ 找不到 800G_TRX 資料，Equipment Performance 將跳過")
        return

    module = trx_result["module"]
    equipment_map = trx_result.get("equipment_map", {})
    df = trx_result["df"].copy()
    df = _add_equipment_column(df, module, equipment_map)
    location_column = find_location_column(list(df.columns))
    df = _add_location_column(df, location_column)
    component_column = module.find_component_column(list(df.columns))
    sort_columns: list[str] = []
    if hasattr(module, "determine_sort_columns"):
        sort_columns = module.determine_sort_columns(df)

    for sheet_name in [EQUIPMENT_PERFORMANCE_SHEET, EQUIPMENT_PERFORMANCE_DATA_SHEET]:
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
    data_ws = workbook.create_sheet(EQUIPMENT_PERFORMANCE_DATA_SHEET)
    data_ws.sheet_state = "hidden"
    chart_ws = workbook.create_sheet(EQUIPMENT_PERFORMANCE_SHEET)

    data_next_row = 1
    chart_next_row = 1
    data_next_row, chart_next_row = populate_equipment_performance_section(
        data_ws,
        chart_ws,
        df,
        categories=["DDMI"],
        items=EQUIPMENT_PERFORMANCE_DDMI_ITEMS,
        data_start_row=data_next_row,
        chart_start_row=chart_next_row,
        section_title="800G_TRX DDMI",
        group_with_location=False,
        component_column=component_column,
        sort_columns=sort_columns,
    )
    data_next_row, chart_next_row = populate_equipment_performance_section(
        data_ws,
        chart_ws,
        df,
        categories=["ATS"],
        items=EQUIPMENT_PERFORMANCE_ATS_ITEMS,
        data_start_row=data_next_row,
        chart_start_row=chart_next_row,
        section_title="800G_TRX ATS",
        group_with_location=False,
        component_column=component_column,
        sort_columns=sort_columns,
    )
    populate_equipment_performance_section(
        data_ws,
        chart_ws,
        df,
        categories=["LT", "HT", "RT"],
        items=EQUIPMENT_PERFORMANCE_TH_ITEMS,
        data_start_row=data_next_row,
        chart_start_row=chart_next_row,
        section_title="800G_TRX LT/HT/RT",
        group_with_location=True,
        split_by_location=True,
        component_column=component_column,
        sort_columns=sort_columns,
    )


def main() -> None:
    args = parse_args()
    start_date = parse_date(args.start_date).isoformat()
    end_date = parse_date(args.end_date).isoformat()

    base_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = normalize_output_path(base_dir, args.output_path, args.start_date, args.end_date)
    out_path = ensure_unique_output_path(out_path)

    reports = [
        {
            "sheet_prefix": "800G_TRX",
            "module_file": "800G_TRX_TEST.py",
            "module_name": "report_trx_test",
            "categories": ["ATS", "DDMI", "LT", "HT", "RT", "其他"],
            "category_builder": lambda df, module: module.classify_ch_number_series(df[module.CH_NUMBER_HEADER]),
        },
        {
            "sheet_prefix": "800G_Fixed_BER",
            "module_file": "800G_Fixed_BER_Test.py",
            "module_name": "report_fixed_ber_test",
            "categories": ["3T_BER", "其他"],
            "category_builder": lambda df, module: df[module.CH_NUMBER_HEADER].apply(module.classify_ch_number),
        },
        {
            "sheet_prefix": "BER_Symbol_Error",
            "module_file": "BER_Symbol_Error_Test.py",
            "module_name": "report_symbol_error_test",
            "categories": ["TC_BER", "其他"],
            "category_builder": lambda df, module: df[module.CH_NUMBER_HEADER].apply(module.classify_ch_number),
        },
    ]

    workbook = load_template_workbook(base_dir)
    ensure_data_analysis_template(workbook)
    report_results: dict[str, dict[str, object]] = {}

    try:
        for report in reports:
            module_path = os.path.join(base_dir, report["module_file"])
            module = load_module(module_path, report["module_name"])
            result = build_report(
                module,
                report["sheet_prefix"],
                report["categories"],
                report["category_builder"],
                workbook,
                start_date,
                end_date,
                base_dir,
            )
            report_results[report["sheet_prefix"]] = result

        populate_combined_data_analysis(workbook, report_results)
        populate_equipment_performance_sheet(workbook, report_results)
        equipment_status = build_equipment_status_table(report_results)
        write_dataframe_to_sheet(workbook, EQUIPMENT_STATUS_SHEET, equipment_status)
        format_equipment_status_sheet(workbook[EQUIPMENT_STATUS_SHEET])
        move_sheet_after(workbook, EQUIPMENT_STATUS_SHEET, DATA_ANALYSIS_SHEET)
        move_sheet_after(workbook, EQUIPMENT_PERFORMANCE_SHEET, EQUIPMENT_STATUS_SHEET)
        hide_sheet(workbook, "Error Code")
        hide_sheet(workbook, EQUIPMENT_PERFORMANCE_DATA_SHEET)
        hide_sheet(workbook, "800G_TRX 其他")
        hide_sheet(workbook, "800G_Fixed_BER 其他")
        hide_sheet(workbook, "BER_Symbol_Error 其他")

        workbook.save(out_path)
        print("📁 Combined Excel 已輸出：", out_path)

    except Exception as exc:
        print("❌ 查詢或匯出失敗：")
        print(exc)
        sys.exit(2)


if __name__ == "__main__":
    main()
